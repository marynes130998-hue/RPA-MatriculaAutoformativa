from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from services.db_service import execute_query_df
from config.settings import EMAIL_DESTINATARIOS, EMAIL_CONFIG
from email.mime.base import MIMEBase
from email import encoders
from db import queries
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import pandas as pd
import os
import base64
from pathlib import Path
from openpyxl.utils import get_column_letter

SCOPES = ["https://www.googleapis.com/auth/gmail.send"]

def get_gmail_service():

    creds = None

    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                "enviarcorreo.json",
                SCOPES
            )
            creds = flow.run_local_server(port=0)

        with open("token.json", "w") as token:
            token.write(creds.to_json())

    service = build("gmail", "v1", credentials=creds)

    return service

def obtener_destinatarios(oferta, grupo):
    df = pd.read_excel("recursos/destinatarios_correo.xlsx")
    row = df[(df["oferta"] == oferta) & (df["grupo"] == grupo)]
    recipients = row["destinatarios"].values[0]

    return recipients.split(";")

def attach_file(message, filepath):

    part = MIMEBase("application", "octet-stream")

    with open(filepath, "rb") as f:
        part.set_payload(f.read())

    encoders.encode_base64(part)

    filename = os.path.basename(filepath)

    part.add_header(
        "Content-Disposition",
        f'attachment; filename="{filename}"'
    )

    message.attach(part)

def build_email_body(total):
    table_cfg = EMAIL_CONFIG["table"]

    tabla_html = f"""
<table style="border-collapse: collapse; font-family: Arial;">
    <tr>
        <th style="border:1px solid #999; padding:8px; text-align:left;">
            {table_cfg["header_1"]}
        </th>
        <th style="border:1px solid #999; padding:8px;">
            {table_cfg["header_2"]}
        </th>
    </tr>
    <tr style="background-color:{table_cfg["row_color"]}; font-weight:bold;">
        <td style="border:1px solid #999; padding:8px;">
            {table_cfg["total_label"]}
        </td>
        <td style="border:1px solid #999; padding:8px; text-align:right;">
            {total}
        </td>
    </tr>
</table>
"""
    signature_html = "<br>".join(EMAIL_CONFIG["signature"])

    body = f"""
<p>{EMAIL_CONFIG["subject_text"]}</p>

<p>{EMAIL_CONFIG["intro_text"]}</p>

{tabla_html}

<p>{EMAIL_CONFIG["footer_text"]}</p>

<p>{signature_html}</p>
"""

    return body

def construir_reporte(nombre_oferta, nombre_grupo, course_id, tipo_oferta, id_oferta):

    if tipo_oferta == "CURSO":
        df = execute_query_df(queries.QUERY_INFO_PARTICIPANTE, (course_id,))
    elif tipo_oferta == "PROGRAMA":
        df = execute_query_df(queries.QUERY_INFO_PARTICIPANTE_PROGRAMA, (id_oferta,))

    # Renombrar columna "usuario_documento" como "DNI"
    df = df.rename(columns={
    "usuario_documento": "DNI"
    })

    df["OBSERVACION TI"] = "MATRICULADO"

    # 🔹 Conteos
    total_registros = len(df)

    # 🔹 1. Generar fecha formato 17.03.26
    fecha = datetime.now().strftime("%d.%m.%y")

    # 🔹 2. Construir nombre del archivo
    if tipo_oferta == "CURSO":
        nombre_archivo = f"Reporte Matriculados Detalle {nombre_grupo}_{nombre_oferta}_{fecha}.xlsx"
    elif tipo_oferta == "PROGRAMA":
        nombre_archivo = f"Reporte Matriculados Detalle {nombre_oferta}_{fecha}.xlsx"

    # 🔹 3. Ruta carpeta reportes
    carpeta_reportes = os.path.join(os.getcwd(), "reportes")

    # Crear carpeta si no existe
    os.makedirs(carpeta_reportes, exist_ok=True)

    ruta_completa = os.path.join(carpeta_reportes, nombre_archivo)

    # 🔹 4. Guardar Excel
    with pd.ExcelWriter(ruta_completa, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Matricula")
        worksheet = writer.sheets["Matricula"]

        # Formato cabecera
        header_format = writer.book.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'middle',
            'align': 'center',
            'fg_color': '#1F4E78',
            'font_color': 'white',
            'border': 1
        })

        # Aplicar formato a encabezados
        for col_num, col_name in enumerate(df.columns):
            if col_name == "OBSERVACION TI":
                format_custom = writer.book.add_format({
                    'bold': True,
                    'align': 'center',
                    'fg_color': '#FFC000',
                    'border': 1
                })
            else:
                format_custom = header_format

            worksheet.write(0, col_num, col_name, format_custom)

        # Ajustar ancho columnas
        for i, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(col)
            )
            worksheet.set_column(i, i, min(max_length + 2, 40))

    # 🔹 5. Retornar ruta
    return ruta_completa, total_registros

def send_email_exito(registro):
    service = get_gmail_service()

    message = MIMEMultipart()

    nombre_oferta = registro["nombre_oferta"]
    nombre_grupo = registro["grupo"]
    course_id = registro["course_id"]
    tipo_oferta = registro["tipo_oferta"].upper()
    id_oferta = registro["id_oferta"]

    ruta_completa, total = construir_reporte(nombre_oferta, nombre_grupo, course_id, tipo_oferta, id_oferta)
    body = build_email_body(total)
    to_emails = obtener_destinatarios(nombre_oferta, nombre_grupo)
    message["to"] = ", ".join(to_emails)

    if tipo_oferta == "CURSO":
        message["subject"] = f"Reporte de Matrícula del {nombre_grupo} del curso {nombre_oferta}"
    elif tipo_oferta == "PROGRAMA":
         message["subject"] = f"Reporte de Matrícula del programa {nombre_oferta}"

    message.attach(MIMEText(body, "html"))

    attach_file(message, ruta_completa)

    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()

    service.users().messages().send(
        userId="me",
        body={"raw": raw}
    ).execute()

def send_email_error(error_info):
    service = get_gmail_service()

    message = MIMEMultipart()

    to_emails = EMAIL_DESTINATARIOS
    message["to"] = ", ".join(to_emails)

    message["subject"] = f"ROBOT MATRICULA AUTOFORMATIVA | ERROR | {error_info['categoria']}"

    body = f"""
    <p>Estimados,</p>

    <p>Se produjo un error en la ejecución:</p>

    <p>Categoria: {error_info['categoria']}</p>
    <p>ID: {error_info['id']}</p>
    <p>Detalle: {error_info['detalle']}</p>
    <p>Trace: {error_info['trace']}</p>

    <p>Saludos cordiales, </p>
    <p>Equipo de Automatización </p>
    <p>DIFODS TI </p>
    """
    message.attach(MIMEText(body, "html"))

    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()

    service.users().messages().send(
        userId="me",
        body={"raw": raw}
    ).execute()

def construir_reporte_final(validaciones):
    # Lista de validaciones
    df = pd.DataFrame(validaciones)

    # 1. Generar timestamp (año-mes-día hora-minuto-segundo)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 2. Definir carpeta "reportes"
    ruta_reportes = Path("reportes")
    ruta_reportes.mkdir(exist_ok=True)  # crea la carpeta si no existe

    # 3. Nombre del archivo
    nombre_archivo = f"Reporte Validaciones {timestamp}.xlsx"

    # 4. Ruta completa
    ruta_completa = ruta_reportes / nombre_archivo

    # 5. Crear Excel con nombre de hoja
    with pd.ExcelWriter(ruta_completa, engine="openpyxl") as writer:
        sheet_name = "Validaciones"
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Obtener worksheet
        ws = writer.sheets[sheet_name]

        # 🔹 Autofit columnas
        for col_idx, col in enumerate(df.columns, 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)

            # considerar encabezado
            max_length = max(max_length, len(col))

            # considerar datos
            for cell in ws[col_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            # ajustar ancho (+2 de margen)
            ws.column_dimensions[col_letter].width = max_length + 2
    
    return ruta_completa

def send_email_info(validaciones):
    service = get_gmail_service()

    message = MIMEMultipart()

    ruta_completa = construir_reporte_final(validaciones)

    to_emails = EMAIL_DESTINATARIOS
    message["to"] = ", ".join(to_emails)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    message["subject"] = f"ROBOT MATRICULA AUTOFORMATIVA | INFO | RESULTADO DE LA EJECUCIÓN DEL DÍA {timestamp}"

    body = f"""
    <p>Estimados,</p>

    <p>Se adjunta resultados de la ejecución del robot.</p>

    <p>Saludos cordiales, </p>
    <p>Equipo de Automatización </p>
    <p>DIFODS TI </p>
    """

    message.attach(MIMEText(body, "html"))

    attach_file(message, ruta_completa)

    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()

    service.users().messages().send(
        userId="me",
        body={"raw": raw}
    ).execute()

def send_email_no_data():
    service = get_gmail_service()

    message = MIMEMultipart()

    to_emails = EMAIL_DESTINATARIOS
    message["to"] = ", ".join(to_emails)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    message["subject"] = f"ROBOT MATRICULA AUTOFORMATIVA | INFO | NO HAY DATA PARA LA EJECUCIÓN DEL DÍA {timestamp}"

    body = f"""
    <p>Estimados,</p>

    <p>No se encontró data para ejecutar la matrícula. Revisar base de datos. </p>

    <p>Saludos cordiales, </p>
    <p>Equipo de Automatización </p>
    <p>DIFODS TI </p>
    """

    message.attach(MIMEText(body, "html"))

    raw = base64.urlsafe_b64encode(message.as_bytes()).decode()

    service.users().messages().send(
        userId="me",
        body={"raw": raw}
    ).execute()